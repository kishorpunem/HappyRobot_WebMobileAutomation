import openpyxl
import os
from datetime import datetime

# folder_path = r"C:\Users\kisho\PycharmProjects\FrameWork_Playwright\ExcelSheet"
# file_path = os.path.join(folder_path, "CreateInquiryData.xlsx")
file_path = os.path.abspath(r"C:\Users\kisho\PycharmProjects\FrameWork_Playwright\ExcelSheet\CreateInquiry.xlsx")

def save_inquiry_to_excel(inquiry_number, country="UAE"):
    print(f"📁 Saving Inquiry to Excel: {file_path}")

    # Create file if not exists
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "InquiryData"
        # Headers
        sheet["A1"] = "TestScenario"
        sheet["B1"] = "InquiryNumber"
        sheet["C1"] = "Country"
        sheet["D1"] = "CreatedDateTime"
    else:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

    # Always store latest inquiry in row 2
    sheet["A2"] = "CreateInquiry"
    sheet["B2"] = inquiry_number
    sheet["C2"] = country
    sheet["D2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb.save(file_path)
    print(f"✅ Inquiry saved: {inquiry_number}")


def read_inquiry_from_excel():
    if not os.path.exists(file_path):
        print("❌ Excel file missing!")
        return None

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    return sheet["B2"].value
