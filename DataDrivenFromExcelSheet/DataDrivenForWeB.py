import openpyxl

class CreateinquiryExecutionData:
    @staticmethod
    def getTestdata(CreateInquiryData):
        book = openpyxl.load_workbook(
            "C:\\Users\\kisho\\PycharmProjects\\FrameWork_Playwright\\ExcelSheet\\CreateInquiryData.xlsx")

        if CreateInquiryData not in book.sheetnames:
            raise ValueError(
                f"❌ Sheet '{CreateInquiryData}' not found. Available sheets: {book.sheetnames}"
            )

        sheet = book[CreateInquiryData]
        headers = [cell.value for cell in sheet[1]]
        data = [
            dict(zip(headers, row))
            for row in sheet.iter_rows(min_row=2, values_only=True)
        ]
        return data

#
# if __name__ == "__main__":
#     data = CreateinquiryExecutionData.getTestdata("CreateInquiryData")
#     print("✅ Excel data read successfully:\n")
#     for record in data:
#         print(record)



# def get_inquiry_test_data():
#     file_path = r"C:\Users\kisho\PycharmProjects\FrameWork_Playwright\ExcelSheet\testdata.xlsx"
#     workbook = openpyxl.load_workbook(file_path)
#     sheet = workbook.active
#
#     headers = [str(cell.value).strip().lower() for cell in sheet[1]]  # normalize headers
#     test_data = []
#
#     for row in sheet.iter_rows(min_row=2, values_only=True):
#         if all(cell is None for cell in row):
#             continue
#         data_dict = dict(zip(headers, row))
#         test_data.append(data_dict)
#
#     return test_data

# ✅ Temporary Debug Section
# if __name__ == "__main__":
#     data = get_inquiry_test_data()
#     print("✅ Excel data read successfully:")
#     for record in data:
#         print(record)